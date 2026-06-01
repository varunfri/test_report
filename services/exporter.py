import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import Union, IO
from services.logger import logger

class DataExporter:
    @staticmethod
    def generate_pivot_statistics(df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate pivot table statistics: Region vs Status count.
        """
        logger.info("Generating pivot summary statistics...")
        if df.empty:
            logger.warning("Dataframe is empty. Returning basic statistics template.")
            return pd.DataFrame(columns=["Region"])
        
        # Create a pivot table: index is Region, columns are unique Statuses
        pivot = df.pivot_table(
            index="Region",
            columns="Testcase Status",
            values="Testcase ID",
            aggfunc="count",
            fill_value=0
        )
        
        # Reset index to make 'Region' a regular column
        pivot = pivot.reset_index()
        
        # Add a Totals row at the bottom
        numeric_cols = [c for c in pivot.columns if c != "Region"]
        totals = {"Region": "Total"}
        for col in numeric_cols:
            totals[col] = pivot[col].sum()
            
        pivot = pd.concat([pivot, pd.DataFrame([totals])], ignore_index=True)
        if not pivot.empty:
            sl_no = list(range(1, len(pivot))) + [""]
            pivot.insert(0, "Sl.No.", sl_no)
        logger.info(f"Summary pivot generated. Regions counted: {len(pivot) - 1}")
        return pivot

    @classmethod
    def export_to_excel(cls, df: pd.DataFrame) -> bytes:
        """
        Export consolidated DataFrame into styled regional sheets and a summary pivot.
        """
        logger.info(f"Starting Excel workbook export with {len(df)} rows...")
        # Create statistics sheet
        stats_df = cls.generate_pivot_statistics(df)
        
        # Extract unique regions for separate sheets
        regions = []
        if not df.empty:
            regions = sorted(df["Region"].unique())
            
        # Use an in-memory bytes buffer
        output = io.BytesIO()
        
        # Write using Pandas ExcelWriter with openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Write the Summary statistics sheet first
            stats_df.to_excel(writer, sheet_name="Statistics", index=False)
            
            # 1.1 Write NA of All regions sheet containing all regions
            df_na = df[df["Testcase Status"] == "NA"].copy() if not df.empty else pd.DataFrame(columns=df.columns)
            if not df_na.empty:
                df_na.insert(0, "Sl.No.", range(1, len(df_na) + 1))
            else:
                df_na.insert(0, "Sl.No.", [])
            df_na.to_excel(writer, sheet_name="NA of All regions", index=False)
            
            # 1.2 Write blocked of all regions sheet containing all regions
            df_blocked = df[df["Testcase Status"] == "Blocked"].copy() if not df.empty else pd.DataFrame(columns=df.columns)
            if not df_blocked.empty:
                df_blocked.insert(0, "Sl.No.", range(1, len(df_blocked) + 1))
            else:
                df_blocked.insert(0, "Sl.No.", [])
            df_blocked.to_excel(writer, sheet_name="blocked of all regions", index=False)
            
            # 2. Write each region's data into its own separate sheet
            for region in regions:
                df_region = df[df["Region"] == region].copy()
                if not df_region.empty:
                    df_region.insert(0, "Sl.No.", range(1, len(df_region) + 1))
                else:
                    df_region.insert(0, "Sl.No.", [])
                df_region.to_excel(writer, sheet_name=str(region), index=False)
            
            # Access workbook and sheets to apply premium styling
            workbook = writer.book
            
            # Define standard styles
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Segoe UI", size=10)
            total_font = Font(name="Segoe UI", size=10, bold=True)
            
            # Distinct dynamic status styling
            blocked_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # soft red
            blocked_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B") # dark red
            
            na_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # soft blue
            na_font = Font(name="Segoe UI", size=10, bold=True, color="1E40AF") # dark blue
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            
            double_bottom_border = Border(
                top=Side(style='thin', color='475569'),
                bottom=Side(style='double', color='475569')
            )
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            # Formatting loop
            ws_names = ["Statistics", "NA of All regions", "blocked of all regions"] + [str(reg) for reg in regions]
            for ws_name in ws_names:
                ws = workbook[ws_name]
                
                # Make sure gridlines are visible
                ws.views.sheetView[0].showGridLines = True
                
                # Format headers (Row 1)
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
                
                # Format data cells
                is_stats_sheet = (ws_name == "Statistics")
                for row_idx in range(2, ws.max_row + 1):
                    is_total_row = (is_stats_sheet and row_idx == ws.max_row)
                    
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        if is_total_row:
                            cell.font = total_font
                            cell.border = double_bottom_border
                            if col_idx == 2:
                                cell.alignment = align_left
                            else:
                                cell.alignment = align_center
                        else:
                            cell.font = data_font
                            cell.border = thin_border
                            
                            # Alignment and dynamic cell styling logic
                            if not is_stats_sheet:
                                # Align text-heavy and ID columns to left, metadata and status to center
                                # Index mapping: 1=Sl.No, 2=Region, 3=Module, 4=Function, 5=Testcase ID, 6=Tester, 7=Testcase Status, 8=Comment
                                if col_idx in [3, 4, 5, 6, 8]:
                                    cell.alignment = align_left
                                else:
                                    cell.alignment = align_center
                                    
                                # Apply status colors dynamically for Testcase Status (Column 7)
                                if col_idx == 7:
                                    val_str = str(cell.value).strip()
                                    if val_str == "Blocked":
                                        cell.fill = blocked_fill
                                        cell.font = blocked_font
                                    elif val_str == "NA":
                                        cell.fill = na_fill
                                        cell.font = na_font
                            else:
                                # Statistics: Region is left (Column 2), others are center
                                if col_idx == 2:
                                    cell.alignment = align_left
                                else:
                                    cell.alignment = align_center
                
                # Auto-fit columns
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    # Add a padding of 4
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    
        return output.getvalue()
